"""
Responsabilità unica: esportare un DataSource come file Excel (.xlsx).
Separato da excel_import.py per rispettare SRP.
"""
import io

import openpyxl

from crm.models import DataSource, Record


def export_group(data_sources: list) -> bytes:
    """Esporta più DataSource come fogli separati in un unico workbook Excel."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # rimuove il foglio vuoto predefinito

    for ds in data_sources:
        ws = wb.create_sheet(title=(ds.label or ds.name)[:31])  # Excel max 31 chars
        ws.append(ds.columns)
        records = Record.objects.filter(data_source=ds).order_by('-is_favorite', 'id')
        for record in records:
            row = [str(record.data.get(col, '')) if record.data.get(col) is not None else '' for col in ds.columns]
            ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_datasource(data_source: DataSource) -> bytes:
    """Serializza un DataSource come workbook Excel e restituisce i byte raw."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = data_source.label or data_source.name

    columns = data_source.columns
    ws.append(columns)

    records = Record.objects.filter(data_source=data_source).order_by('-is_favorite', 'id')
    for record in records:
        row = [str(record.data.get(col, '')) if record.data.get(col) is not None else '' for col in columns]
        ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
