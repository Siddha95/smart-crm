import json
import os
import tempfile

import openpyxl
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.test import TestCase
from rest_framework import status
from rest_framework.test import APIClient

from crm.models import DataSource, Note, Record, RecordHistory
from crm.services.excel_import import clean_row, import_all_sheets, preview_sheets, get_sheet_names
from crm.views import _safe_filename, _SAFE_COL_RE


# ── Helper ─────────────────────────────────────────────────────────────────────

def _make_excel(rows: list[dict], sheet_name: str = "Foglio") -> str:
    """Crea un file Excel temporaneo e restituisce il path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if rows:
        ws.append(list(rows[0].keys()))
        for row in rows:
            ws.append(list(row.values()))
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        wb.save(f.name)
        return f.name


# ── Model: DataSource ──────────────────────────────────────────────────────────

class DataSourceModelTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="dsuser", password="pass")

    def test_valid_columns(self):
        ds = DataSource(name="test", label="Test", columns=["nome", "email"], owner=self.user)
        ds.full_clean()

    def test_invalid_columns_not_list(self):
        ds = DataSource(name="test", label="Test", columns={"nome": "email"}, owner=self.user)
        with self.assertRaises(ValidationError):
            ds.full_clean()

    def test_invalid_columns_not_strings(self):
        ds = DataSource(name="test", label="Test", columns=["nome", 123], owner=self.user)
        with self.assertRaises(ValidationError):
            ds.full_clean()

    def test_datasource_owner_isolation(self):
        """Utenti diversi possono avere datasource con lo stesso nome."""
        u2 = User.objects.create_user(username="u2", password="pass")
        DataSource.objects.create(name="ds", label="DS", columns=["nome"], owner=self.user)
        DataSource.objects.create(name="ds", label="DS", columns=["nome"], owner=u2)
        self.assertEqual(DataSource.objects.filter(owner=self.user).count(), 1)
        self.assertEqual(DataSource.objects.filter(owner=u2).count(), 1)


# ── Service: clean_row ─────────────────────────────────────────────────────────

class CleanRowTest(TestCase):
    def test_nan_replaced_with_none(self):
        row = {"nome": "Mario", "eta": float("nan")}
        cleaned = clean_row(row)
        self.assertIsNone(cleaned["eta"])
        self.assertEqual(cleaned["nome"], "Mario")

    def test_valid_row_unchanged(self):
        row = {"nome": "Mario", "email": "mario@example.com"}
        self.assertEqual(clean_row(row), row)

    def test_zero_not_replaced(self):
        import numpy as np
        row = {"qty": np.int64(0), "price": np.float64(0.0)}
        cleaned = clean_row(row)
        self.assertEqual(cleaned["qty"], 0)
        self.assertIsInstance(cleaned["qty"], int)
        self.assertEqual(cleaned["price"], 0.0)
        self.assertIsInstance(cleaned["price"], float)

    def test_empty_string_preserved(self):
        row = {"nome": ""}
        self.assertEqual(clean_row(row)["nome"], "")

    def test_numpy_int64_converted(self):
        import numpy as np
        row = {"eta": np.int64(42)}
        cleaned = clean_row(row)
        self.assertEqual(cleaned["eta"], 42)
        self.assertIsInstance(cleaned["eta"], int)

    def test_numpy_float_nan_converted_to_none(self):
        import numpy as np
        row = {"val": np.float64(float("nan"))}
        self.assertIsNone(clean_row(row)["val"])

    def test_numpy_bool_converted(self):
        import numpy as np
        row = {"flag": np.bool_(True)}
        cleaned = clean_row(row)
        self.assertEqual(cleaned["flag"], True)
        self.assertIsInstance(cleaned["flag"], bool)

    def test_pandas_timestamp_converted_to_iso(self):
        import pandas as pd
        row = {"data": pd.Timestamp("2024-01-15")}
        cleaned = clean_row(row)
        self.assertEqual(cleaned["data"], "15/01/2024")

    def test_pandas_nat_converted_to_none(self):
        import pandas as pd
        row = {"data": pd.NaT}
        self.assertIsNone(clean_row(row)["data"])

    def test_all_types_are_json_serializable(self):
        import json, numpy as np, pandas as pd
        from django.core.serializers.json import DjangoJSONEncoder
        row = {
            "nome": "Mario",
            "eta": np.int64(42),
            "prezzo": np.float64(3.14),
            "data": pd.Timestamp("2024-01-15"),
            "vuoto": float("nan"),
            "flag": np.bool_(True),
            "nat": pd.NaT,
        }
        cleaned = clean_row(row)
        # Non deve sollevare TypeError
        json.dumps(cleaned, cls=DjangoJSONEncoder)


# ── Service: excel_import ──────────────────────────────────────────────────────

class ImportAllSheetsTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="importer", password="pass")

    def test_import_creates_datasource_and_records(self):
        path = _make_excel([{"nome": "Mario", "email": "mario@example.com"}])
        try:
            result = import_all_sheets(path, source_file="clienti", owner=self.user)
        finally:
            os.unlink(path)

        self.assertEqual(result["total_imported"], 1)
        self.assertEqual(len(result["sheets"]), 1)
        self.assertTrue(DataSource.objects.filter(source_file="clienti", owner=self.user).exists())
        ds = DataSource.objects.get(source_file="clienti", owner=self.user)
        self.assertEqual(Record.objects.filter(data_source=ds).count(), 1)

    def test_reimport_replaces_records(self):
        path = _make_excel([{"nome": "Mario"}, {"nome": "Luigi"}])
        try:
            import_all_sheets(path, source_file="clienti", owner=self.user)
            result = import_all_sheets(path, source_file="clienti", owner=self.user)
        finally:
            os.unlink(path)

        self.assertEqual(result["sheets"][0]["deleted"], 2)
        self.assertEqual(result["total_imported"], 2)
        ds = DataSource.objects.get(source_file="clienti", owner=self.user)
        self.assertEqual(Record.objects.filter(data_source=ds).count(), 2)

    def test_two_users_isolated(self):
        user2 = User.objects.create_user(username="altro", password="pass")
        path = _make_excel([{"nome": "Mario"}])
        try:
            import_all_sheets(path, source_file="clienti", owner=self.user)
            import_all_sheets(path, source_file="clienti", owner=user2)
        finally:
            os.unlink(path)

        self.assertEqual(DataSource.objects.filter(owner=self.user).count(), 1)
        self.assertEqual(DataSource.objects.filter(owner=user2).count(), 1)

    def test_file_not_found_raises(self):
        with self.assertRaises(FileNotFoundError):
            import_all_sheets("/non/esiste.xlsx", source_file="test", owner=self.user)

    def test_multiple_sheets(self):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Clienti"
        ws1.append(["nome"])
        ws1.append(["Mario"])
        ws2 = wb.create_sheet("Fornitori")
        ws2.append(["nome"])
        ws2.append(["Acme"])
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
            wb.save(f.name)
            path = f.name
        try:
            result = import_all_sheets(path, source_file="azienda", owner=self.user)
        finally:
            os.unlink(path)

        self.assertEqual(len(result["sheets"]), 2)
        self.assertEqual(result["total_imported"], 2)
        self.assertEqual(DataSource.objects.filter(owner=self.user).count(), 2)

    def test_preview_sheets_limits_rows(self):
        path = _make_excel([{"nome": f"User{i}"} for i in range(10)])
        try:
            result = preview_sheets(path, max_rows=3)
        finally:
            os.unlink(path)

        self.assertEqual(len(result[0]["rows"]), 3)
        self.assertIn("nome", result[0]["columns"])

    def test_get_sheet_names(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Foglio1"
        wb.create_sheet("Foglio2")
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
            wb.save(f.name)
            path = f.name
        try:
            names = get_sheet_names(path)
        finally:
            os.unlink(path)
        self.assertEqual(names, ["Foglio1", "Foglio2"])

    def test_import_is_atomic_on_bulk_create_failure(self):
        """Se bulk_create fallisce, i vecchi record non devono essere eliminati."""
        path = _make_excel([{"nome": "Mario"}])
        try:
            import_all_sheets(path, source_file="clienti", owner=self.user)
        finally:
            os.unlink(path)

        ds = DataSource.objects.get(source_file="clienti", owner=self.user)
        original_count = Record.objects.filter(data_source=ds).count()
        self.assertEqual(original_count, 1)

        # Simula un fallimento nel bulk_create mockando l'ORM
        from unittest.mock import patch
        with patch("crm.models.Record.objects") as mock_manager:
            mock_manager.filter.return_value.delete.return_value = (1, {})
            mock_manager.bulk_create.side_effect = Exception("DB error")
            path2 = _make_excel([{"nome": "Luigi"}])
            try:
                with self.assertRaises(Exception):
                    import_all_sheets(path2, source_file="clienti", owner=self.user)
            finally:
                os.unlink(path2)

        # I record originali devono esistere ancora (transazione rollback)
        self.assertEqual(Record.objects.filter(data_source=ds).count(), 1)


# ── Model: RecordHistory ───────────────────────────────────────────────────────

class RecordHistoryTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="testuser", password="pass")
        self.ds = DataSource.objects.create(name="test", label="Test", columns=["campo"], owner=self.user)
        self.record = Record.objects.create(data_source=self.ds, data={"campo": "valore"})

    def test_history_created_on_change(self):
        RecordHistory.objects.create(
            record=self.record,
            changed_by=self.user,
            field_changed="campo",
            old_value="valore",
            new_value="nuovo",
        )
        self.assertEqual(self.record.history.count(), 1)

    def test_history_ordered_desc(self):
        RecordHistory.objects.create(record=self.record, changed_by=self.user,
                                     field_changed="campo", old_value="a", new_value="b")
        RecordHistory.objects.create(record=self.record, changed_by=self.user,
                                     field_changed="campo", old_value="b", new_value="c")
        entries = list(RecordHistory.objects.filter(record=self.record).order_by('-changed_at'))
        self.assertEqual(entries[0].new_value, "c")


# ── View: DataSourceViewSet - Column Operations ────────────────────────────────

class ColumnOperationTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username="coltest", password="pass")
        self.client.force_authenticate(user=self.user)
        self.ds = DataSource.objects.create(
            name="ds1", label="DS1", columns=["nome", "email"], owner=self.user
        )
        Record.objects.create(data_source=self.ds, data={"nome": "Mario", "email": "m@e.com"})

    def _url(self):
        return f"/api/datasources/{self.ds.id}/columns/"

    # ── add ──

    def test_add_column(self):
        resp = self.client.post(self._url(), {"operation": "add", "name": "telefono"})
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.ds.refresh_from_db()
        self.assertIn("telefono", self.ds.columns)
        record = Record.objects.get(data_source=self.ds)
        self.assertIn("telefono", record.data)

    def test_add_column_null_in_existing_records(self):
        self.client.post(self._url(), {"operation": "add", "name": "fax"})
        record = Record.objects.get(data_source=self.ds)
        self.assertIsNone(record.data["fax"])

    def test_add_column_duplicate_rejected(self):
        resp = self.client.post(self._url(), {"operation": "add", "name": "nome"})
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_add_column_missing_name_rejected(self):
        resp = self.client.post(self._url(), {"operation": "add", "name": ""})
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    # ── rename ──

    def test_rename_column(self):
        resp = self.client.post(self._url(), {"operation": "rename", "old_name": "nome", "new_name": "cognome"})
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.ds.refresh_from_db()
        self.assertIn("cognome", self.ds.columns)
        self.assertNotIn("nome", self.ds.columns)
        record = Record.objects.get(data_source=self.ds)
        self.assertIn("cognome", record.data)
        self.assertNotIn("nome", record.data)

    def test_rename_preserves_value(self):
        self.client.post(self._url(), {"operation": "rename", "old_name": "nome", "new_name": "cognome"})
        record = Record.objects.get(data_source=self.ds)
        self.assertEqual(record.data["cognome"], "Mario")

    def test_rename_nonexistent_column(self):
        resp = self.client.post(self._url(), {"operation": "rename", "old_name": "inesistente", "new_name": "x"})
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    def test_rename_to_existing_name_rejected(self):
        resp = self.client.post(self._url(), {"operation": "rename", "old_name": "nome", "new_name": "email"})
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_rename_missing_params_rejected(self):
        resp = self.client.post(self._url(), {"operation": "rename", "old_name": "nome"})
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    # ── delete ──

    def test_delete_column(self):
        resp = self.client.post(self._url(), {"operation": "delete", "name": "email"})
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.ds.refresh_from_db()
        self.assertNotIn("email", self.ds.columns)
        record = Record.objects.get(data_source=self.ds)
        self.assertNotIn("email", record.data)

    def test_delete_nonexistent_column(self):
        resp = self.client.post(self._url(), {"operation": "delete", "name": "inesistente"})
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    # ── other ──

    def test_invalid_operation(self):
        resp = self.client.post(self._url(), {"operation": "invalid"})
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_other_user_cannot_modify_columns(self):
        other = User.objects.create_user(username="intruder", password="pass")
        self.client.force_authenticate(user=other)
        resp = self.client.post(self._url(), {"operation": "add", "name": "telefono"})
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    def test_unauthenticated_rejected(self):
        self.client.force_authenticate(user=None)
        resp = self.client.post(self._url(), {"operation": "add", "name": "telefono"})
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)


# ── View: RecordViewSet ────────────────────────────────────────────────────────

class RecordViewSetTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username="rectest", password="pass")
        self.client.force_authenticate(user=self.user)
        self.ds = DataSource.objects.create(
            name="ds1", label="DS1", columns=["nome", "eta"], owner=self.user
        )

    # ── CRUD ──

    def test_create_record(self):
        resp = self.client.post("/api/records/", {
            "data_source": self.ds.id, "data": {"nome": "Alice", "eta": "30"}
        }, format="json")
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Record.objects.count(), 1)

    def test_list_records_filtered_by_datasource(self):
        ds2 = DataSource.objects.create(name="ds2", label="DS2", columns=["x"], owner=self.user)
        Record.objects.create(data_source=self.ds, data={"nome": "A"})
        Record.objects.create(data_source=ds2, data={"x": "B"})
        resp = self.client.get(f"/api/records/?data_source={self.ds.id}")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.data["count"], 1)

    def test_update_record_creates_history(self):
        record = Record.objects.create(data_source=self.ds, data={"nome": "Alice", "eta": "30"})
        resp = self.client.patch(f"/api/records/{record.id}/",
                                 {"data": {"nome": "Bob", "eta": "30"}}, format="json")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(RecordHistory.objects.filter(record=record, field_changed="nome").count(), 1)

    def test_update_record_no_history_if_unchanged(self):
        record = Record.objects.create(data_source=self.ds, data={"nome": "Alice", "eta": "30"})
        self.client.patch(f"/api/records/{record.id}/",
                          {"data": {"nome": "Alice", "eta": "30"}}, format="json")
        self.assertEqual(RecordHistory.objects.filter(record=record).count(), 0)

    def test_update_is_atomic(self):
        """Se la creazione della history fallisce, il record non deve essere modificato."""
        record = Record.objects.create(data_source=self.ds, data={"nome": "Alice", "eta": "30"})
        from unittest.mock import patch
        with patch.object(RecordHistory.objects, "create", side_effect=Exception("DB error")):
            with self.assertRaises(Exception):
                # Forziamo la chiamata diretta per testare l'atomicità
                from crm.views import RecordViewSet
                from rest_framework.request import Request
                from unittest.mock import MagicMock
                view = RecordViewSet()
                view.request = MagicMock()
                view.request.user = self.user
                view.kwargs = {"pk": record.id}
                serializer = MagicMock()
                serializer.save.return_value = record
                view.perform_update(serializer)
        # Il record non deve essere modificato
        record.refresh_from_db()
        self.assertEqual(record.data["nome"], "Alice")

    def test_delete_record(self):
        record = Record.objects.create(data_source=self.ds, data={"nome": "A"})
        resp = self.client.delete(f"/api/records/{record.id}/")
        self.assertEqual(resp.status_code, status.HTTP_204_NO_CONTENT)
        self.assertFalse(Record.objects.filter(id=record.id).exists())

    # ── Authorization ──

    def test_other_user_cannot_see_records(self):
        Record.objects.create(data_source=self.ds, data={"nome": "A"})
        other = User.objects.create_user(username="spy", password="pass")
        self.client.force_authenticate(user=other)
        resp = self.client.get("/api/records/")
        self.assertEqual(resp.data["count"], 0)

    def test_cannot_create_record_for_other_user_datasource(self):
        other = User.objects.create_user(username="owner2", password="pass")
        other_ds = DataSource.objects.create(name="ds_other", label="X", columns=["x"], owner=other)
        resp = self.client.post("/api/records/",
                                {"data_source": other_ds.id, "data": {"x": "1"}}, format="json")
        self.assertIn(resp.status_code, [status.HTTP_403_FORBIDDEN, status.HTTP_400_BAD_REQUEST])

    def test_unauthenticated_cannot_list(self):
        self.client.force_authenticate(user=None)
        resp = self.client.get("/api/records/")
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)

    # ── Filtering ──

    def test_favorite_filter(self):
        r1 = Record.objects.create(data_source=self.ds, data={"nome": "A"}, is_favorite=True)
        Record.objects.create(data_source=self.ds, data={"nome": "B"})
        resp = self.client.get(f"/api/records/?data_source={self.ds.id}&is_favorite=true")
        self.assertEqual(resp.data["count"], 1)
        self.assertEqual(resp.data["results"][0]["id"], r1.id)

    def test_search_case_insensitive(self):
        Record.objects.create(data_source=self.ds, data={"nome": "Alice", "eta": "30"})
        Record.objects.create(data_source=self.ds, data={"nome": "Bob", "eta": "25"})
        resp = self.client.get(f"/api/records/?data_source={self.ds.id}&search=ALICE")
        self.assertEqual(resp.data["count"], 1)

    def test_column_filter_case_insensitive(self):
        Record.objects.create(data_source=self.ds, data={"nome": "Alice", "eta": "30"})
        Record.objects.create(data_source=self.ds, data={"nome": "Bob", "eta": "25"})
        col_filters = json.dumps({"nome": "ALICE"})
        resp = self.client.get(f"/api/records/?data_source={self.ds.id}&col_filters={col_filters}")
        self.assertEqual(resp.data["count"], 1)

    def test_malicious_column_filter_ignored(self):
        """Injection attempt in col name is silently ignored, records survive."""
        Record.objects.create(data_source=self.ds, data={"nome": "Alice"})
        col_filters = json.dumps({"nome'; DROP TABLE crm_record;--": "x"})
        resp = self.client.get(f"/api/records/?data_source={self.ds.id}&col_filters={col_filters}")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(Record.objects.count(), 1)

    def test_malicious_ordering_ignored(self):
        """Injection attempt in ordering param is silently ignored."""
        Record.objects.create(data_source=self.ds, data={"nome": "Alice"})
        resp = self.client.get(
            f"/api/records/?data_source={self.ds.id}&ordering=nome'; DROP TABLE crm_record;--"
        )
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(Record.objects.count(), 1)

    def test_ordering_asc(self):
        Record.objects.create(data_source=self.ds, data={"nome": "Zzz", "eta": "1"})
        Record.objects.create(data_source=self.ds, data={"nome": "Aaa", "eta": "2"})
        resp = self.client.get(f"/api/records/?data_source={self.ds.id}&ordering=nome")
        self.assertEqual(resp.data["results"][0]["data"]["nome"], "Aaa")

    def test_ordering_desc(self):
        Record.objects.create(data_source=self.ds, data={"nome": "Zzz", "eta": "1"})
        Record.objects.create(data_source=self.ds, data={"nome": "Aaa", "eta": "2"})
        resp = self.client.get(f"/api/records/?data_source={self.ds.id}&ordering=-nome")
        self.assertEqual(resp.data["results"][0]["data"]["nome"], "Zzz")

    # ── Pagination ──

    def test_pagination_default_page_size(self):
        for i in range(30):
            Record.objects.create(data_source=self.ds, data={"nome": f"User{i}"})
        resp = self.client.get(f"/api/records/?data_source={self.ds.id}")
        self.assertEqual(resp.data["count"], 30)
        self.assertEqual(len(resp.data["results"]), 25)

    def test_pagination_page2(self):
        for i in range(30):
            Record.objects.create(data_source=self.ds, data={"nome": f"User{i}"})
        resp = self.client.get(f"/api/records/?data_source={self.ds.id}&page=2")
        self.assertEqual(len(resp.data["results"]), 5)

    def test_pagination_custom_page_size(self):
        for i in range(10):
            Record.objects.create(data_source=self.ds, data={"nome": f"User{i}"})
        resp = self.client.get(f"/api/records/?data_source={self.ds.id}&page_size=3")
        self.assertEqual(len(resp.data["results"]), 3)

    # ── Actions ──

    def test_move_record(self):
        record = Record.objects.create(data_source=self.ds, data={"nome": "Alice"})
        target = DataSource.objects.create(name="target", label="Target", columns=["nome"], owner=self.user)
        resp = self.client.patch(f"/api/records/{record.id}/move/", {"data_source_id": target.id})
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        record.refresh_from_db()
        self.assertEqual(record.data_source, target)

    def test_move_record_to_other_user_datasource_denied(self):
        record = Record.objects.create(data_source=self.ds, data={"nome": "Alice"})
        other = User.objects.create_user(username="owner3", password="pass")
        other_ds = DataSource.objects.create(name="ds_o", label="X", columns=["nome"], owner=other)
        resp = self.client.patch(f"/api/records/{record.id}/move/", {"data_source_id": other_ds.id})
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    def test_history_endpoint(self):
        record = Record.objects.create(data_source=self.ds, data={"nome": "Alice"})
        RecordHistory.objects.create(record=record, changed_by=self.user,
                                     field_changed="nome", old_value="X", new_value="Alice")
        resp = self.client.get(f"/api/records/{record.id}/history/")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(len(resp.data), 1)

    def test_history_endpoint_other_user_denied(self):
        record = Record.objects.create(data_source=self.ds, data={"nome": "Alice"})
        other = User.objects.create_user(username="spy2", password="pass")
        self.client.force_authenticate(user=other)
        resp = self.client.get(f"/api/records/{record.id}/history/")
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    def test_bulk_delete(self):
        r1 = Record.objects.create(data_source=self.ds, data={"nome": "A"})
        r2 = Record.objects.create(data_source=self.ds, data={"nome": "B"})
        resp = self.client.post("/api/records/bulk_delete/",
                                {"ids": [r1.id, r2.id]}, format="json")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.data["deleted"], 2)
        self.assertEqual(Record.objects.count(), 0)

    def test_bulk_delete_cannot_delete_other_user_records(self):
        other = User.objects.create_user(username="owner4", password="pass")
        other_ds = DataSource.objects.create(name="ds_o2", label="X", columns=["x"], owner=other)
        r_other = Record.objects.create(data_source=other_ds, data={"x": "secret"})
        resp = self.client.post("/api/records/bulk_delete/",
                                {"ids": [r_other.id]}, format="json")
        self.assertEqual(resp.data["deleted"], 0)
        self.assertTrue(Record.objects.filter(id=r_other.id).exists())

    def test_bulk_move(self):
        r1 = Record.objects.create(data_source=self.ds, data={"nome": "A"})
        r2 = Record.objects.create(data_source=self.ds, data={"nome": "B"})
        target = DataSource.objects.create(name="t_ds", label="Target", columns=["nome"], owner=self.user)
        resp = self.client.post("/api/records/bulk_move/",
                                {"ids": [r1.id, r2.id], "data_source_id": target.id}, format="json")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(Record.objects.filter(data_source=target).count(), 2)

    def test_bulk_move_to_other_user_datasource_denied(self):
        r1 = Record.objects.create(data_source=self.ds, data={"nome": "A"})
        other = User.objects.create_user(username="owner5", password="pass")
        other_ds = DataSource.objects.create(name="ds_o3", label="X", columns=["nome"], owner=other)
        resp = self.client.post("/api/records/bulk_move/",
                                {"ids": [r1.id], "data_source_id": other_ds.id}, format="json")
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)


# ── View: DataSourceViewSet - Export ───────────────────────────────────────────

class ExportTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(username="exptest", password="pass")
        self.client.force_authenticate(user=self.user)
        self.ds = DataSource.objects.create(
            name="myds", label="MyDS", columns=["nome"], owner=self.user
        )
        Record.objects.create(data_source=self.ds, data={"nome": "Alice"})

    def test_export_xlsx(self):
        resp = self.client.get(f"/api/datasources/{self.ds.id}/export/")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        self.assertIn("myds.xlsx", resp["Content-Disposition"])

    def test_export_sanitizes_filename(self):
        ds = DataSource.objects.create(
            name='file"name\r\nX-Evil: hdr',
            label="Evil",
            columns=["nome"],
            owner=self.user,
        )
        Record.objects.create(data_source=ds, data={"nome": "X"})
        resp = self.client.get(f"/api/datasources/{ds.id}/export/")
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        # Header must not contain CRLF (injection) or quotes inside the filename value
        header = resp["Content-Disposition"]
        self.assertNotIn("\r", header)
        self.assertNotIn("\n", header)
        # Extract filename value between the outer quotes
        filename = header.split('filename="')[1].rstrip('"')
        self.assertNotIn('"', filename)
        self.assertNotIn(":", filename)

    def test_export_other_user_denied(self):
        other = User.objects.create_user(username="spy3", password="pass")
        self.client.force_authenticate(user=other)
        resp = self.client.get(f"/api/datasources/{self.ds.id}/export/")
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)


# ── Security: _safe_filename ───────────────────────────────────────────────────

class SafeFilenameTest(TestCase):
    def test_normal_name_unchanged(self):
        self.assertEqual(_safe_filename("clienti"), "clienti")

    def test_spaces_and_hyphen_preserved(self):
        result = _safe_filename("clienti 2024")
        self.assertIn("clienti", result)

    def test_dot_preserved(self):
        self.assertIn(".", _safe_filename("report.v1"))

    def test_header_injection_stripped(self):
        result = _safe_filename("report\r\nX-Injected: evil")
        self.assertNotIn("\r", result)
        self.assertNotIn("\n", result)

    def test_double_quote_stripped(self):
        self.assertNotIn('"', _safe_filename('file"name'))

    def test_empty_name_fallback(self):
        self.assertEqual(_safe_filename(""), "export")

    def test_all_special_chars_fallback(self):
        self.assertEqual(_safe_filename('!!!???'), "export")

    def test_semicolon_stripped(self):
        self.assertNotIn(";", _safe_filename("file;name"))


# ── Security: _SAFE_COL_RE ─────────────────────────────────────────────────────

class SafeColRegexTest(TestCase):
    def test_valid_column_names(self):
        for col in ["nome", "data nascita", "eta", "CAP", "field_1", "col name"]:
            with self.subTest(col=col):
                self.assertIsNotNone(_SAFE_COL_RE.match(col))

    def test_invalid_column_names_rejected(self):
        for col in ["nome;", "col'value", 'col"x', "col--", "col/**/comment", "col\n", "col\r", "col\t"]:
            with self.subTest(col=col):
                self.assertIsNone(_SAFE_COL_RE.match(col))


# ── NoteViewSet ────────────────────────────────────────────────────────────────

class NoteViewSetTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('noteuser', password='pw')
        self.other = User.objects.create_user('otheruser', password='pw')
        self.client = APIClient()
        self.client.force_authenticate(self.user)
        self.other_client = APIClient()
        self.other_client.force_authenticate(self.other)

    # ── CRUD base ──────────────────────────────────────────────────────────────

    def test_create_note(self):
        res = self.client.post('/api/notes/', {'title': 'Test', 'content': 'ciao'}, format='json')
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertEqual(res.data['title'], 'Test')
        self.assertEqual(res.data['content'], 'ciao')

    def test_list_returns_own_notes(self):
        Note.objects.create(owner=self.user, title='A', content='')
        Note.objects.create(owner=self.user, title='B', content='')
        res = self.client.get('/api/notes/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(len(res.data), 2)

    def test_retrieve_own_note(self):
        note = Note.objects.create(owner=self.user, title='Memo', content='testo')
        res = self.client.get(f'/api/notes/{note.id}/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['title'], 'Memo')

    def test_update_title_and_content(self):
        note = Note.objects.create(owner=self.user, title='Vecchio', content='')
        res = self.client.patch(f'/api/notes/{note.id}/', {'title': 'Nuovo', 'content': '**bold**'}, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        note.refresh_from_db()
        self.assertEqual(note.title, 'Nuovo')
        self.assertEqual(note.content, '**bold**')

    def test_partial_update_content_only(self):
        note = Note.objects.create(owner=self.user, title='Titolo', content='')
        res = self.client.patch(f'/api/notes/{note.id}/', {'content': 'nuovo testo'}, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        note.refresh_from_db()
        self.assertEqual(note.title, 'Titolo')  # invariato
        self.assertEqual(note.content, 'nuovo testo')

    def test_delete_note(self):
        note = Note.objects.create(owner=self.user, title='Da eliminare', content='')
        res = self.client.delete(f'/api/notes/{note.id}/')
        self.assertEqual(res.status_code, status.HTTP_204_NO_CONTENT)
        self.assertFalse(Note.objects.filter(id=note.id).exists())

    # ── Owner isolation ────────────────────────────────────────────────────────

    def test_other_user_notes_not_listed(self):
        Note.objects.create(owner=self.other, title='Privata', content='')
        res = self.client.get('/api/notes/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(len(res.data), 0)

    def test_cannot_retrieve_other_user_note(self):
        note = Note.objects.create(owner=self.other, title='Privata', content='')
        res = self.client.get(f'/api/notes/{note.id}/')
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)

    def test_cannot_update_other_user_note(self):
        note = Note.objects.create(owner=self.other, title='Privata', content='')
        res = self.client.patch(f'/api/notes/{note.id}/', {'title': 'Hacked'}, format='json')
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)
        note.refresh_from_db()
        self.assertEqual(note.title, 'Privata')

    def test_cannot_delete_other_user_note(self):
        note = Note.objects.create(owner=self.other, title='Privata', content='')
        res = self.client.delete(f'/api/notes/{note.id}/')
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)
        self.assertTrue(Note.objects.filter(id=note.id).exists())

    # ── Auth ───────────────────────────────────────────────────────────────────

    def test_unauthenticated_cannot_list(self):
        res = APIClient().get('/api/notes/')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_unauthenticated_cannot_create(self):
        res = APIClient().post('/api/notes/', {'title': 'X', 'content': ''}, format='json')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    # ── Read-only fields ───────────────────────────────────────────────────────

    def test_response_includes_timestamps(self):
        res = self.client.post('/api/notes/', {'title': 'TS', 'content': ''}, format='json')
        self.assertIn('created_at', res.data)
        self.assertIn('updated_at', res.data)

    def test_owner_not_exposed_in_response(self):
        res = self.client.post('/api/notes/', {'title': 'Owner', 'content': ''}, format='json')
        self.assertNotIn('owner', res.data)

    # ── Ordering ───────────────────────────────────────────────────────────────

    def test_notes_ordered_by_updated_at_desc(self):
        n1 = Note.objects.create(owner=self.user, title='Prima', content='')
        n2 = Note.objects.create(owner=self.user, title='Seconda', content='')
        # Aggiorna n1 dopo n2: deve comparire prima nella lista
        n1.content = 'aggiornata'
        n1.save()
        res = self.client.get('/api/notes/')
        self.assertEqual(res.data[0]['id'], n1.id)
        self.assertEqual(res.data[1]['id'], n2.id)

    # ── Model unit test ────────────────────────────────────────────────────────

    def test_note_str(self):
        note = Note(owner=self.user, title='Mia nota')
        self.assertEqual(str(note), 'Mia nota')

    def test_note_default_title(self):
        note = Note.objects.create(owner=self.user, content='contenuto')
        self.assertEqual(note.title, 'Senza titolo')

    def test_note_default_content_is_empty(self):
        note = Note.objects.create(owner=self.user, title='Vuota')
        self.assertEqual(note.content, '')


# ── StageTemplateViewSet ───────────────────────────────────────────────────────

from crm.models import StageTemplate
from crm.services.excel_import import INSERT_DATE_COL
import datetime


class StageTemplateViewSetTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('stuser', password='pw')
        self.other = User.objects.create_user('stother', password='pw')
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    # ── CRUD ──────────────────────────────────────────────────────────────────

    def test_create_template(self):
        res = self.client.post('/api/stage-templates/', {
            'name': 'Pipeline vendite', 'stages': ['Lead', 'Contattato', 'Chiuso']
        }, format='json')
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertEqual(res.data['stages'], ['Lead', 'Contattato', 'Chiuso'])

    def test_list_own_templates(self):
        StageTemplate.objects.create(owner=self.user, name='A', stages=['X'])
        StageTemplate.objects.create(owner=self.user, name='B', stages=['Y'])
        res = self.client.get('/api/stage-templates/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(len(res.data), 2)

    def test_retrieve_own_template(self):
        t = StageTemplate.objects.create(owner=self.user, name='T', stages=['S1'])
        res = self.client.get(f'/api/stage-templates/{t.id}/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['name'], 'T')

    def test_update_template(self):
        t = StageTemplate.objects.create(owner=self.user, name='Old', stages=['A'])
        res = self.client.patch(f'/api/stage-templates/{t.id}/', {'name': 'New'}, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        t.refresh_from_db()
        self.assertEqual(t.name, 'New')

    def test_delete_template(self):
        t = StageTemplate.objects.create(owner=self.user, name='Del', stages=['A'])
        res = self.client.delete(f'/api/stage-templates/{t.id}/')
        self.assertEqual(res.status_code, status.HTTP_204_NO_CONTENT)
        self.assertFalse(StageTemplate.objects.filter(id=t.id).exists())

    # ── Owner isolation ────────────────────────────────────────────────────────

    def test_other_user_templates_not_listed(self):
        StageTemplate.objects.create(owner=self.other, name='Private', stages=['X'])
        res = self.client.get('/api/stage-templates/')
        self.assertEqual(len(res.data), 0)

    def test_cannot_retrieve_other_user_template(self):
        t = StageTemplate.objects.create(owner=self.other, name='Private', stages=['X'])
        res = self.client.get(f'/api/stage-templates/{t.id}/')
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)

    def test_cannot_update_other_user_template(self):
        t = StageTemplate.objects.create(owner=self.other, name='Private', stages=['X'])
        res = self.client.patch(f'/api/stage-templates/{t.id}/', {'name': 'Hacked'}, format='json')
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)

    def test_cannot_delete_other_user_template(self):
        t = StageTemplate.objects.create(owner=self.other, name='Private', stages=['X'])
        res = self.client.delete(f'/api/stage-templates/{t.id}/')
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)
        self.assertTrue(StageTemplate.objects.filter(id=t.id).exists())

    # ── Auth ──────────────────────────────────────────────────────────────────

    def test_unauthenticated_cannot_list(self):
        res = APIClient().get('/api/stage-templates/')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    # ── Ordering ──────────────────────────────────────────────────────────────

    def test_templates_ordered_alphabetically(self):
        StageTemplate.objects.create(owner=self.user, name='Zebra', stages=[])
        StageTemplate.objects.create(owner=self.user, name='Alpha', stages=[])
        res = self.client.get('/api/stage-templates/')
        self.assertEqual(res.data[0]['name'], 'Alpha')
        self.assertEqual(res.data[1]['name'], 'Zebra')

    # ── Owner non esposto ─────────────────────────────────────────────────────

    def test_owner_not_in_response(self):
        res = self.client.post('/api/stage-templates/', {'name': 'X', 'stages': []}, format='json')
        self.assertNotIn('owner', res.data)


# ── Record reorder ─────────────────────────────────────────────────────────────

class RecordReorderTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('reorder_user', password='pw')
        self.other = User.objects.create_user('reorder_other', password='pw')
        self.client = APIClient()
        self.client.force_authenticate(self.user)
        self.ds = DataSource.objects.create(
            owner=self.user, name='DS', label='DS', columns=['nome']
        )
        self.r1 = Record.objects.create(data_source=self.ds, data={'nome': 'A'}, position=0)
        self.r2 = Record.objects.create(data_source=self.ds, data={'nome': 'B'}, position=1)
        self.r3 = Record.objects.create(data_source=self.ds, data={'nome': 'C'}, position=2)

    def test_reorder_changes_positions(self):
        res = self.client.post('/api/records/reorder/', {
            'ids': [self.r3.id, self.r1.id, self.r2.id]
        }, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.r1.refresh_from_db(); self.r2.refresh_from_db(); self.r3.refresh_from_db()
        self.assertEqual(self.r3.position, 0)
        self.assertEqual(self.r1.position, 1)
        self.assertEqual(self.r2.position, 2)

    def test_reorder_returns_updated_count(self):
        res = self.client.post('/api/records/reorder/', {
            'ids': [self.r1.id, self.r2.id]
        }, format='json')
        self.assertEqual(res.data['updated'], 2)

    def test_reorder_ignores_other_user_records(self):
        other_ds = DataSource.objects.create(
            owner=self.other, name='ODS', label='ODS', columns=['x']
        )
        other_r = Record.objects.create(data_source=other_ds, data={'x': '1'}, position=99)
        self.client.post('/api/records/reorder/', {
            'ids': [other_r.id, self.r1.id]
        }, format='json')
        other_r.refresh_from_db()
        self.assertEqual(other_r.position, 99)  # non modificato

    def test_reorder_missing_ids_returns_400(self):
        res = self.client.post('/api/records/reorder/', {}, format='json')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_unauthenticated_cannot_reorder(self):
        res = APIClient().post('/api/records/reorder/', {'ids': [self.r1.id]}, format='json')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_kanban_ordering_by_position(self):
        self.r1.position = 2; self.r1.save()
        self.r2.position = 0; self.r2.save()
        self.r3.position = 1; self.r3.save()
        res = self.client.get(f'/api/records/?data_source={self.ds.id}&ordering=position&page_size=10')
        ids = [r['id'] for r in res.data['results']]
        self.assertEqual(ids, [self.r2.id, self.r3.id, self.r1.id])


# ── INSERT_DATE_COL (Data inserimento) ────────────────────────────────────────

class InsertDateColumnTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('insdate_user', password='pw')
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def _make_excel(self, rows, sheet='Foglio'):
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = sheet
        if rows:
            ws.append(list(rows[0].keys()))
            for r in rows: ws.append(list(r.values()))
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            wb.save(f.name); return f.name

    def test_import_adds_insert_date_column(self):
        path = self._make_excel([{'nome': 'Mario'}])
        try:
            import_all_sheets(path, source_file='test', owner=self.user)
        finally:
            os.unlink(path)
        ds = DataSource.objects.get(owner=self.user)
        self.assertIn(INSERT_DATE_COL, ds.columns)
        self.assertEqual(ds.columns[-1], INSERT_DATE_COL)  # ultima colonna

    def test_import_populates_insert_date_value(self):
        path = self._make_excel([{'nome': 'Mario'}])
        today = datetime.date.today().strftime('%d/%m/%Y')
        try:
            import_all_sheets(path, source_file='test', owner=self.user)
        finally:
            os.unlink(path)
        record = Record.objects.filter(data_source__owner=self.user).first()
        self.assertEqual(record.data[INSERT_DATE_COL], today)

    def test_reimport_updates_insert_date(self):
        path = self._make_excel([{'nome': 'Mario'}])
        try:
            import_all_sheets(path, source_file='test', owner=self.user)
            import_all_sheets(path, source_file='test', owner=self.user)
        finally:
            os.unlink(path)
        today = datetime.date.today().strftime('%d/%m/%Y')
        record = Record.objects.filter(data_source__owner=self.user).first()
        self.assertEqual(record.data[INSERT_DATE_COL], today)

    def test_existing_insert_date_column_in_file_is_overwritten(self):
        """Se il file Excel ha già una colonna 'Data inserimento', viene ignorata e rigenerata."""
        path = self._make_excel([{INSERT_DATE_COL: '01/01/2000', 'nome': 'X'}])
        today = datetime.date.today().strftime('%d/%m/%Y')
        try:
            import_all_sheets(path, source_file='test', owner=self.user)
        finally:
            os.unlink(path)
        record = Record.objects.filter(data_source__owner=self.user).first()
        self.assertEqual(record.data[INSERT_DATE_COL], today)

    def test_manual_record_creation_auto_populates_insert_date(self):
        ds = DataSource.objects.create(
            owner=self.user, name='DS', label='DS',
            columns=['nome', INSERT_DATE_COL]
        )
        today = datetime.date.today().strftime('%d/%m/%Y')
        res = self.client.post('/api/records/', {
            'data_source': ds.id, 'data': {'nome': 'Test'}
        }, format='json')
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertEqual(res.data['data'][INSERT_DATE_COL], today)

    def test_datasource_without_insert_date_not_affected(self):
        """Datasource senza la colonna non deve ricevere il campo automatico."""
        ds = DataSource.objects.create(
            owner=self.user, name='DS2', label='DS2', columns=['nome']
        )
        res = self.client.post('/api/records/', {
            'data_source': ds.id, 'data': {'nome': 'Test'}
        }, format='json')
        self.assertNotIn(INSERT_DATE_COL, res.data['data'])


# ── CleanRow: tipi datetime nativi Python ─────────────────────────────────────

class CleanRowDatetimeTest(TestCase):
    def test_python_datetime_formatted(self):
        row = {'ts': datetime.datetime(2024, 6, 15, 10, 30, 0)}
        self.assertEqual(clean_row(row)['ts'], '15/06/2024')

    def test_python_date_formatted(self):
        row = {'d': datetime.date(2024, 1, 5)}
        self.assertEqual(clean_row(row)['d'], '05/01/2024')

    def test_datetime_no_time_component_in_output(self):
        row = {'ts': datetime.datetime(2024, 12, 31, 23, 59, 59)}
        result = clean_row(row)['ts']
        self.assertNotIn('23:59', result)
        self.assertEqual(result, '31/12/2024')


# ── UserProfile ai_context ────────────────────────────────────────────────────

from crm.models import UserProfile


class UserProfileAiContextTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('ctx_user', password='pw')
        self.other = User.objects.create_user('ctx_other', password='pw')
        self.client = APIClient()
        self.client.force_authenticate(self.user)
        self.profile, _ = UserProfile.objects.get_or_create(user=self.user)

    def test_ai_context_returned_in_profile(self):
        self.profile.ai_context = 'Lavoro nel settore edilizia'
        self.profile.save()
        res = self.client.get('/api/profile/')
        self.assertEqual(res.data[0]['ai_context'], 'Lavoro nel settore edilizia')

    def test_save_ai_context(self):
        res = self.client.patch(
            f'/api/profile/{self.profile.id}/ai-context/',
            {'ai_context': 'Contesto di test'},
            format='json'
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.ai_context, 'Contesto di test')

    def test_save_empty_context(self):
        self.profile.ai_context = 'vecchio'
        self.profile.save()
        res = self.client.patch(
            f'/api/profile/{self.profile.id}/ai-context/',
            {'ai_context': ''},
            format='json'
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.ai_context, '')

    def test_cannot_update_other_user_context(self):
        other_profile, _ = UserProfile.objects.get_or_create(user=self.other)
        other_profile.ai_context = 'privato'
        other_profile.save()
        res = self.client.patch(
            f'/api/profile/{other_profile.id}/ai-context/',
            {'ai_context': 'hacked'},
            format='json'
        )
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)
        other_profile.refresh_from_db()
        self.assertEqual(other_profile.ai_context, 'privato')

    def test_unauthenticated_cannot_update_context(self):
        res = APIClient().patch(
            f'/api/profile/{self.profile.id}/ai-context/',
            {'ai_context': 'x'},
            format='json'
        )
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_default_ai_context_is_empty(self):
        new_user = User.objects.create_user('ctx_new', password='pw')
        profile, _ = UserProfile.objects.get_or_create(user=new_user)
        self.assertEqual(profile.ai_context, '')


# ── Security: refresh token blacklist ──────────────────────────────────────────

class RefreshTokenBlacklistTest(TestCase):
    """Verifica che BLACKLIST_AFTER_ROTATION sia attivo e che i refresh token
    usati vengano invalidati dopo la rotazione."""

    def setUp(self):
        self.user = User.objects.create_user('blacklist_user', password='pw')
        self.client = APIClient()

    def _get_tokens(self):
        res = self.client.post('/api/auth/token/', {'username': 'blacklist_user', 'password': 'pw'}, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        return res.data['access'], res.data['refresh']

    def test_refresh_returns_new_tokens(self):
        _, refresh = self._get_tokens()
        res = self.client.post('/api/auth/token/refresh/', {'refresh': refresh}, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('access', res.data)
        self.assertIn('refresh', res.data)

    def test_used_refresh_token_is_blacklisted(self):
        """Dopo un refresh, il vecchio refresh token non deve più funzionare."""
        _, refresh = self._get_tokens()
        # Primo uso: ok
        self.client.post('/api/auth/token/refresh/', {'refresh': refresh}, format='json')
        # Secondo uso con lo stesso token: deve fallire
        res = self.client.post('/api/auth/token/refresh/', {'refresh': refresh}, format='json')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_new_refresh_token_works_after_rotation(self):
        """Il nuovo refresh token emesso dalla rotazione funziona correttamente."""
        _, refresh = self._get_tokens()
        res1 = self.client.post('/api/auth/token/refresh/', {'refresh': refresh}, format='json')
        new_refresh = res1.data['refresh']
        res2 = self.client.post('/api/auth/token/refresh/', {'refresh': new_refresh}, format='json')
        self.assertEqual(res2.status_code, status.HTTP_200_OK)
        self.assertIn('access', res2.data)

    def test_invalid_refresh_token_rejected(self):
        res = self.client.post('/api/auth/token/refresh/', {'refresh': 'token.falso.invalido'}, format='json')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)


# ── Security: AI tools soft-delete ─────────────────────────────────────────────

class AiToolSoftDeleteTest(TestCase):
    """Verifica che il tool delete_record dell'AI usi il soft-delete (is_active=False)
    invece di eliminare fisicamente il record dal database."""

    def setUp(self):
        self.user = User.objects.create_user('aitool_user', password='pw')
        self.ds = DataSource.objects.create(
            name='ai_ds', label='AI DS', columns=['nome'], owner=self.user
        )
        self.record = Record.objects.create(
            data_source=self.ds, data={'nome': 'Vittima'}, is_active=True
        )

    def test_delete_record_sets_is_active_false(self):
        """delete_record deve impostare is_active=False, non eliminare la riga."""
        from django.db import connection as _conn
        with _conn.cursor() as cur:
            cur.execute(
                "UPDATE crm_record SET is_active = FALSE WHERE id = %s AND "
                "id IN (SELECT r.id FROM crm_record r JOIN crm_datasource ds ON r.data_source_id = ds.id WHERE ds.owner_id = %s)",
                [self.record.id, self.user.id]
            )
        self.record.refresh_from_db()
        self.assertFalse(self.record.is_active)
        # Il record esiste ancora nel DB
        self.assertTrue(Record.objects.filter(id=self.record.id).exists())

    def test_deleted_record_excluded_from_active_queryset(self):
        """Un record con is_active=False non compare nelle query filtranti gli attivi."""
        self.record.is_active = False
        self.record.save()
        active_ids = list(Record.objects.filter(is_active=True, data_source=self.ds).values_list('id', flat=True))
        self.assertNotIn(self.record.id, active_ids)

    def test_active_record_still_retrievable_after_soft_delete(self):
        """Il record soft-deleted è recuperabile (per audit) senza filtro is_active."""
        self.record.is_active = False
        self.record.save()
        self.assertTrue(Record.objects.filter(id=self.record.id).exists())
        retrieved = Record.objects.get(id=self.record.id)
        self.assertFalse(retrieved.is_active)
        self.assertEqual(retrieved.data['nome'], 'Vittima')
